import os
import tempfile
from openpyxl import Workbook
import subprocess
from concurrent.futures import ProcessPoolExecutor
import time

def get_folder_info(root):
    folder_size = 0
    file_count = 0
    subfolder_size = 0
    subfolder_count = 0

    for folder_path, _, files in os.walk(root):
        for file in files:
            file_path = os.path.join(folder_path, file)
            folder_size += os.path.getsize(file_path)
            file_count += 1

        for subfolder in os.listdir(folder_path):
            subfolder_path = os.path.join(folder_path, subfolder)
            if os.path.isdir(subfolder_path):
                subfolder_count += 1
                subfolder_size += sum(os.path.getsize(os.path.join(subfolder_path, f)) for f in os.listdir(subfolder_path) if os.path.isfile(os.path.join(subfolder_path, f)))

    return {
        'folder': root,
        'file_count': file_count,
        'folder_size': folder_size,
        'subfolder_count': subfolder_count,
        'subfolder_size': subfolder_size
    }

def bytes_to_mb(bytes_size):
    return bytes_size / (1024 * 1024)

def create_excel_chart(data, output_path):
    workbook = Workbook()
    sheet = workbook.active

    sheet['A1'] = 'Folder'
    sheet['B1'] = 'Number of Files'
    sheet['C1'] = 'Total Size (bytes)'
    sheet['D1'] = 'Total Size (MB)'
    sheet['E1'] = 'Number of Subfolders'
    sheet['F1'] = 'Subfolder Size (bytes)'
    sheet['G1'] = 'Subfolder Size (MB)'

    max_depth = max(info['folder'].count(os.sep) for info in data)
    for depth in range(1, max_depth + 1):
        sheet.cell(row=1, column=7 + depth, value=f'Subfolder {depth}')

    for row_num, info in enumerate(data, start=2):
        sheet.cell(row=row_num, column=1, value=info['folder'])
        sheet.cell(row=row_num, column=2, value=info['file_count'])
        sheet.cell(row=row_num, column=3, value=info['folder_size'])
        sheet.cell(row=row_num, column=4, value=bytes_to_mb(info['folder_size']))
        sheet.cell(row=row_num, column=5, value=info['subfolder_count'])
        sheet.cell(row=row_num, column=6, value=info['subfolder_size'])

        folder_parts = info['folder'].split(os.sep)
        for depth, part in enumerate(folder_parts, start=1):
            sheet.cell(row=row_num, column=7 + depth, value=part)

    workbook.save(output_path)

if __name__ == "__main__":
    folder_path = input("Enter the folder path: ")

    start_time = time.time()

    try:
        with ProcessPoolExecutor() as executor:
            futures = [executor.submit(get_folder_info, root) for root, _, _ in os.walk(folder_path)]

        data = [future.result() for future in futures]

        # Add information for the main folder itself
        main_folder_info = {
            'folder': folder_path,
            'file_count': len([f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]),
            'folder_size': sum(os.path.getsize(os.path.join(folder_path, file)) for file in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, file))),
            'subfolder_count': len([f for f in os.listdir(folder_path) if os.path.isdir(os.path.join(folder_path, f))]),
            'subfolder_size': sum(os.path.getsize(os.path.join(folder_path, file)) for file in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, file))),
        }
        data.insert(0, main_folder_info)

        # Get the temporary directory
        temp_dir = tempfile.gettempdir()

        # Create a temporary Excel file path
        temp_excel_path = os.path.join(temp_dir, 'folder_info.xlsx')

        create_excel_chart(data, temp_excel_path)

        print(f"Data written to {temp_excel_path}")

        # Open the Excel file
        subprocess.Popen(['start', 'excel', temp_excel_path], shell=True)

    except PermissionError:
        print("Permission denied. Make sure you have read permissions for the specified folder.")

    end_time = time.time()
    total_time = end_time - start_time
    print(f"Total time taken: {total_time} seconds")
